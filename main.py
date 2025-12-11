import sys
import math
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from core.models import LoadInput, InstallationParams, ConduitType, ConductorMaterial, InsulationRating
from core.converters import convert_power_unit, convert_length_unit
from standards.nec_logic import NECLogic
from standards.nec_tables import NEC_310_16_COPPER

def get_installation_params():
    print("\n--- Parámetros de Instalación (Condiciones Ambientales) ---")
    
    # Ambient Temp - Ask more clearly? No, standard is fine.
    try:
        temp = float(input("Temperatura Ambiente (°C) [Default 30]: ") or 30.0)
    except ValueError:
        temp = 30.0
        
    # Duct Type
    print("Tipos de Ducto: (1) PVC, (2) Acero (Steel), (3) Aluminio")
    d_choice = input("Seleccione Ducto [1]: ")
    if d_choice == "2":
        conduit = ConduitType.STEEL
    elif d_choice == "3":
        conduit = ConduitType.ALUMINUM
    else:
        conduit = ConduitType.PVC
        
    return temp, conduit

def get_loads_input(temp, conduit):
    loads = []
    print("\n--- Gestión de Cargas ---")
    
    while True:
        print(f"\n[Carga #{len(loads)+1}]")
        name = input("Nombre de la Carga: ").strip()
        if not name: break
        
        try:
            # 1. Quantity
            qty_str = input("Cantidad de cargas idénticas [1]: ").strip()
            quantity = int(qty_str) if qty_str else 1
            
            # 2. Power & Unit
            p_input_str = input("Potencia (ej: 1000 W, 10 KW, 5 HP, 20 A, 50 KVA): ").strip()
            # Split value and unit
            import re
            match = re.match(r"([0-9\.]+)\s*([a-zA-Z]+)", p_input_str)
            if match:
                val = float(match.group(1))
                unit = match.group(2)
            else:
                # Default to Watts if no unit
                val = float(p_input_str)
                unit = "W"
                
            voltage = float(input("Voltaje (V): "))
            phases = int(input("Fases (1 o 3): "))
            pf = float(input("Factor de Potencia [0.9]: ") or 0.9)
            
            # Convert Unit
            watts, amps_override = convert_power_unit(val, unit, voltage, phases, pf)
            
            # NEC Flags
            is_motor = input("¿Es Motor? (s/n) [n]: ").lower() == 's'
            is_cont = False
            if not is_motor:
                is_cont = input("¿Es Carga Continua (>3h)? (s/n) [n]: ").lower() == 's'
            
            # 3. Insulation Rating Selection
            print("Seleccione Temperatura Nominal del Conductor:")
            print("1. 75°C (THWN, RH...) - Estándar")
            print("2. 90°C (THHN, THWN-2, XHHW-2)")
            t_choice = input("Opción [1]: ").strip()
            ins_rating = InsulationRating.TEMP_90 if t_choice == "2" else InsulationRating.TEMP_75
            
            # 4. Grouping / Raceway Count
            # "Numero total de conductores para realizar las correcciones"
            # Default to 3 (just this circuit).
            # If user puts many in one pipe, they specify here.
            try:
                count = int(input(f"N° Total de conductores en el ducto (NEC 310.15(C)(1)) [{3 if phases==3 else 2}]: ") or (3 if phases==3 else 2))
            except ValueError:
                count = 3
                
            # 5. Length & Unit
            l_input_str = input("Longitud del circuito (ej: 50 m, 100 ft): ").strip()
            match_l = re.match(r"([0-9\.]+)\s*([a-zA-Z]+)", l_input_str)
            if match_l:
                l_val = float(match_l.group(1))
                l_unit = match_l.group(2)
            else:
                l_val = float(l_input_str)
                l_unit = "m"
            
            length_m = convert_length_unit(l_val, l_unit)
            
            # Create Objects
            load_obj = LoadInput(
                name=name, power_watts=watts, voltage=voltage, phases=phases,
                is_continuous=is_cont, is_motor=is_motor, power_factor=pf,
                quantity=quantity, override_amps=amps_override
            )
            
            install_params = InstallationParams(
                length_meters=length_m,
                conduit_type=conduit,
                ambient_temp_c=temp,
                raceway_count=count,
                conductor_material=ConductorMaterial.COPPER, 
                insulation_rating=ins_rating
            )
            
            loads.append((load_obj, install_params))
            
        except ValueError as e:
            print(f"Error en entrada de datos: {e}. Intente de nuevo.")
            
        more = input("¿Agregar otra carga? (s/n): ").lower()
        if more != 's':
            break
            
    return loads

def export_to_excel(loads_results, feeder_result):
    wb = Workbook()
    
    # --- Sheet 1: Detalles Cargas ---
    ws1 = wb.active
    ws1.title = "Circuitos Derivados"
    
    headers = ["Cant.", "Carga", "Potencia (W)", "Amps Unit", "Voltaje", "Fases", "F.P.", "Long.(m)", "Temp Rating", "Agrup.", "Calibre", "Ampacidad", "Breaker", "% VD", "Notas"]
    ws1.append(headers)
    
    # Style
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    for item in loads_results:
        load = item["load"]
        res = item["result"]
        params = item["params"]
        
        # Determine Amps to show
        amps_show = load.override_amps if load.override_amps else (load.power_watts / (load.voltage * (1.732 if load.phases==3 else 1) * load.power_factor))
        
        ws1.append([
            load.quantity,
            load.name, 
            f"{load.power_watts:.1f}", 
            f"{amps_show:.2f}",
            load.voltage, load.phases, load.power_factor,
            f"{params.length_meters:.1f}",
            f"{params.insulation_rating.value} C",
            params.raceway_count,
            res.size, res.ampacity, res.breaker_rating,
            f"{res.voltage_drop_percent:.2f}%", res.reference_notes
        ])
        
    for col in ws1.columns:
        ws1.column_dimensions[col[0].column_letter].width = 15

    # --- Sheet 2: Feeder ---
    ws2 = wb.create_sheet("Resumen Alimentador")
    ws2.append(["CÁLCULO ALIMENTADOR PRINCIPAL (NEC)"])
    ws2.append(["Fecha:", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws2.append([])
    ws2.append(["Parámetro", "Valor"])
    
    ws2.append(["Corriente Total Estimada (A)", f"{feeder_result['total_amps']:.2f}"])
    ws2.append(["Conductor Seleccionado", feeder_result['cable_size']])
    ws2.append(["Configuración", feeder_result['description']])
    ws2.append(["Consideraciones", "Incluye 125% Motor Mayor, 125% Continous, Demand Factors."])
    
    # --- Sheet 3: NEC Reference ---
    ws3 = wb.create_sheet("Ref NEC Tables")
    ws3.append(["NEC 310.16 (Cobre 75°C Base)"])
    for s, data in NEC_310_16_COPPER.items():
        ws3.append([s, f"{data[75]} A"])
        
    # --- Sheet 4: NEC 250.122 (Grounding) ---
    ws4 = wb.create_sheet("Ref NEC 250.122")
    ws4.append(["NEC Tabla 250.122 - Puesta a Tierra"])
    ws4.append(["Protección (A)", "Cond. Cu (AWG/kcmil)"])
    
    grounding_table = [
        (15, "14"), (20, "12"), (60, "10"), (100, "8"),
        (200, "6"), (300, "4"), (400, "3"), (500, "2"),
        (600, "1"), (800, "1/0"), (1000, "2/0"), (1200, "3/0"),
        (1600, "4/0"), (2000, "250"), (2500, "350"), (3000, "400"),
        (4000, "500"), (5000, "700"), (6000, "800")
    ]
    for row in grounding_table:
        ws4.append(row)

    filename = f"Memoria_NEC_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    print(f"\n[INFO] Excel generado: {filename}")

def main():
    print("==========================================================")
    print(" CALCULADORA DE TABLEROS (NEC EDITION V2.1 - PHASE 8)")
    print("==========================================================")
    
    # 1. Global Installation Params
    temp, conduit = get_installation_params()
    
    # 2. Add Loads
    load_data = get_loads_input(temp, conduit)
    
    if not load_data:
        print("No se ingresaron cargas.")
        sys.exit()
        
    print("\nCalculando Circuitos Derivados...")
    print("-" * 120)
    print(f"{'Cant':<4} | {'Carga':<15} | {'Amps/Unit':<9} | {'Calibre':<10} | {'Breaker':<8} | {'% VD':<6} | {'Rating':<10} | {'Notas'}")
    print("-" * 120)
    
    results = []
    
    # 3. Calculate Branch Circuits
    for load, params in load_data:
        # Calculate separately
        result = NECLogic.select_conductor_and_breaker(load, params)
        results.append({"load": load, "params": params, "result": result})
        
        warn = " (!)" if result.voltage_drop_percent > 3.0 else ""
        
        # Display Amps (calculated or override)
        if load.override_amps:
            disp_amps = load.override_amps
        else:
             if load.phases == 1:
                disp_amps = load.power_watts / (load.voltage * load.power_factor)
             else:
                disp_amps = load.power_watts / (math.sqrt(3) * load.voltage * load.power_factor)
        
        rating_str = f"{params.insulation_rating.value}C"
        
        print(f"{load.quantity:<4} | {load.name:<15} | {disp_amps:<9.1f} | {result.size:<10} | {result.breaker_rating:<8} | {result.voltage_drop_percent:<6.2f}{warn} | {rating_str:<10} | {result.reference_notes[:30]}...")
        
    print("-" * 120)
    
    # 4. Calculate Main Feeder
    print("\nCalculando Alimentador Principal (NEC 430.24 / 215.2)...")
    only_loads = [x[0] for x in load_data]
    feeder_res = NECLogic.calculate_main_feeder(only_loads)
    
    print(f"Corriente Total Estimada: {feeder_res['total_amps']:.2f} A")
    print(f"Alimentador Sugerido:     {feeder_res['description']}")
    
    # 5. Export
    ask = input("\n¿Exportar reporte a Excel? (s/n): ").lower()
    if ask == 's':
        export_to_excel(results, feeder_res)

if __name__ == "__main__":
    main()
